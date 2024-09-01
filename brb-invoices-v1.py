import os
from openpyxl import Workbook
import pdfplumber
import re
from datetime import datetime

directory = 'invoices'
files = os.listdir(directory)
files_quantity = len(files)

if files_quantity == 0:
    raise Exception("No files found in the directory")

wb = Workbook()
ws = wb.active
ws.title = 'Fatura BRB'

# Definição de colunas
ws['A1'] = 'Data'
ws['B1'] = 'Descrição'
ws['C1'] = 'Valor'


last_empty_line = 1
while ws['A' + str(last_empty_line)].value is not None:
    last_empty_line += 1

for file in files:
    with pdfplumber.open(directory + "/" + file) as pdf:
         pdf_text = ""
         for page in pdf.pages:
            pdf_text += page.extract_text() + "\n"  

    # Pega o que estiver entre os textos 'trasações internacionais' e 'transações nacionais'. Assim consigo ter todas as transações internacionais
    regex_inter = r'Transações internacionais\n(.*?)(?=\nTransações nacionais)'
    match_inter = re.search(regex_inter, pdf_text, re.DOTALL)

    if match_inter:
        international_text = match_inter.group(1)
        inter_transaction_in_list = international_text.splitlines() #Transforma string em lista

        for inter_transaction in inter_transaction_in_list:
            # Expressão regular para extrair data, descrição e valor(Para transações internacionais o valor correto é o segundo valor em real, o primeiro é a cotação do dólar)
            inter_regex = r"(\d{2}/\d{2})\s(.+?)R\$\s[\d,]+.*?R\$\s([\d,]+)"
            inter_match = re.search(inter_regex, inter_transaction)

            if inter_match:
                inter_match_data = inter_match.group(1)
                ws['A{}'.format(last_empty_line)] = inter_match_data

                inter_match_description = inter_match.group(2).strip()
                ws['B{}'.format(last_empty_line)] = inter_match_description

                inter_match_value = inter_match.group(3).replace('.', '').replace(',', '.')
                ws['C{}'.format(last_empty_line)] = float(inter_match_value)

                last_empty_line += 1 


    # Passo 2: Remover a transação internacional que já foi usanda no bloco anterior
    text_without_international_transaction = re.sub(regex_inter, '', pdf_text, flags=re.DOTALL)

    # Regex para pegar todas as trasações nacionais
    regex = r'Transações nacionais\n(.*?)(?:\nEncargos|\Z)'

    # Encontrar o trecho que contém as transações nacionais
    national_transactions = re.search(regex, text_without_international_transaction, re.DOTALL)

    if national_transactions:
        # Extraindo as transações nacionais
        transaction_list = national_transactions.group(1)

        transaction_in_list = transaction_list.splitlines() #Transforma string em lista
        transaction_in_list.pop() #Remove o último elemento da lista que não sera usado

        for transaction in transaction_in_list:
            # Expressão regular para extrair data, descrição e valor
            regex = r"(\d{2}/\d{2})\s(.+?)R\$\s([\d,]+)"
            match = re.search(regex, transaction)

            if match:
                match_data = match.group(1)
                ws['A{}'.format(last_empty_line)] = match_data

                match_description = match.group(2).strip()
                ws['B{}'.format(last_empty_line)] = match_description

                match_value = match.group(3).replace('.', '').replace(',', '.')
                ws['C{}'.format(last_empty_line)] = float(match_value)

                last_empty_line += 1 

    else:
        print("Nenhuma transação nacional encontrada.")


full_now = str(datetime.now()).replace(":", "-")
dot_index = full_now.index(".")
now = full_now[:dot_index]

wb.save("fature - {}.xlsx".format(now))
