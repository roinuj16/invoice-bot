import os
from openpyxl import Workbook
import pdfplumber
import re
from datetime import datetime

def convert_to_float(value_str):
    return float(value_str.replace('.', '').replace(',', '.'))

def process_transaction(transaction, last_empty_line, worksheet, regex):
    match = re.search(regex, transaction)
    if match:
        match_data = match.group(1)
        worksheet[f'A{last_empty_line}'] = match_data

        match_description = match.group(2).strip()
        worksheet[f'B{last_empty_line}'] = match_description

        match_value = convert_to_float(match.group(3))
        worksheet[f'C{last_empty_line}'] = match_value

        return last_empty_line + 1
    return last_empty_line

def process_pdf(file, last_empty_line, worksheet):
    with pdfplumber.open(file) as pdf:
        pdf_text = ""
        for page in pdf.pages:
            pdf_text += page.extract_text() + "\n"

    # Pega o que estiver entre os textos 'trasações internacionais' e 'transações nacionais'. Assim consigo ter todas as transações internacionais
    regex_inter = r'Transações internacionais\n(.*?)(?=\nTransações nacionais)'
    match_inter = re.search(regex_inter, pdf_text, re.DOTALL)

    if match_inter:
        international_text = match_inter.group(1)
        #Transforma string em lista
        inter_transaction_in_list = international_text.splitlines() 
        # Expressão regular para extrair data, descrição e valor(Para transações internacionais o valor correto é o segundo valor em real, o primeiro é a cotação do dólar)
        inter_regex = re.compile(r"(\d{2}/\d{2})\s(.+?)R\$\s[\d,]+.*?R\$\s([\d,]+)")

        for inter_transaction in inter_transaction_in_list:
            last_empty_line = process_transaction(inter_transaction, last_empty_line, worksheet, inter_regex)

    # Passo 2: Remover a transação internacional que já foi usanda no bloco anterior
    text_without_international_transaction = re.sub(regex_inter, '', pdf_text, flags=re.DOTALL)

    # Regex para pegar todas as trasações nacionais
    regex_national = r'Transações nacionais\n(.*?)(?:\nEncargos|\Z)'
    # Encontrar o trecho que contém as transações nacionais
    national_transactions = re.search(regex_national, text_without_international_transaction, re.DOTALL)

    if national_transactions:
        #Transforma string em lista
        transaction_list = national_transactions.group(1).splitlines()
        #Remove o último elemento da lista que não sera usado
        transaction_list.pop()
        # Expressão regular para extrair data, descrição e valor
        national_regex = re.compile(r"(\d{2}/\d{2})\s(.+?)R\$\s([\d,]+)")

        for transaction in transaction_list:
            last_empty_line = process_transaction(transaction, last_empty_line, worksheet, national_regex)

    return last_empty_line

def main():
    directory = 'brb/invoices-pdf'
    files = os.listdir(directory)
    if not files:
        raise Exception("No files found in the directory")

    wb = Workbook()
    ws = wb.active
    ws.title = 'Fatura BRB'

    # Definição de colunas
    ws['A1'] = 'Data'
    ws['B1'] = 'Descrição'
    ws['C1'] = 'Valor'

    last_empty_line = 2
    for file in files:
        last_empty_line = process_pdf(os.path.join(directory, file), last_empty_line, ws)


    now = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    # Diretório onde deseja salvar o arquivo
    save_directory = 'brb/invoices'
    os.makedirs(save_directory, exist_ok=True)  # Criar diretório se não existir

    # Caminho completo para salvar o arquivo
    save_path = os.path.join(save_directory, f"fatura_BRB_{now}.xlsx")
    wb.save(save_path)

   
    # wb.save(f"fatura_BRB_{now}.xlsx")

if __name__ == "__main__":
    main()
